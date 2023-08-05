# -*- coding: utf-8 -*-

import xlrd
import xlwt
from xlutils.copy import copy
from openpyxl import load_workbook



class Excel_deal():
    def __init__(self,src=None):
        self.src=src
    def read_from_excel(self, sheet_inedx=None, sheet_name=None):
        data_gd = []
        data = xlrd.open_workbook(self.src)
        if sheet_inedx != None:
            table = data.sheet_by_index(sheet_inedx)
        if sheet_name != None:
            for x in data.sheet_names():
                if sheet_name in x:
                    new_sheet_name=x
                    break
            table = data.sheet_by_name(new_sheet_name)
        nrows = table.nrows
        for i in range(0, nrows):
            data_gd.append(table.row_values(i))
        return data_gd

    def data_write_to_excel(self,data, file_name, sheet_name):
        df = xlwt.Workbook()
        table = df.add_sheet(sheet_name, cell_overwrite_ok=True)
        for i in range(len(data)):
            temp_data = data[i]
            for j in range(len(temp_data)):
                table.write(i, j, temp_data[j])
        df.save(file_name)
    def write_to_excel(self,msg,h,l,num=0):
        rb=xlrd.open_workbook(self.src,formatting_info=True)
        wb=copy(rb)
        ws=wb.get_sheet(num)
        ws.write(h,l,msg)
        wb.save(self.src)

    def revise_to_excel(self, file_name, sheet_name, data,row=None,column=None):
        wb = load_workbook(file_name)
        ws = wb[sheet_name]
        if row==None and column==None:
            for x in data:
                ws.append(list(x.values()))
        else:
            i=0
            for c in range(*row):
                for l in range(*column):
                    ws.cell(c,l,data[i])
                    i+=1
        wb.save(file_name)
        wb.close()

    def list2dic(self,data,key_num):
        key = data[key_num]
        data_new = []
        for x in data[key_num+1:]:
            data_one = {}
            for i in range(len(x)):
                data_one[key[i]] = x[i]
            data_new.append(data_one)
        return data_new



